from openpyxl import Workbook, load_workbook


async def check_user_id(user_id, sheet):
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] == user_id:
            return True
    return False

async def id_in_excel(user_id, user_name):
    workbook = load_workbook("files\kambo.xlsx")

    sheet = workbook.active

    sheet['A1'] = 'User_id'
    sheet['B1'] = 'User_name'

    workbook.save('files\kambo.xlsx')

    if await check_user_id(user_id, sheet) == False:

        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            row_num += 1

        sheet.cell(row=row_num, column=1).value = user_id
        sheet.cell(row=row_num, column=2).value = f"@{user_name}"

        workbook.save('files\kambo.xlsx')
